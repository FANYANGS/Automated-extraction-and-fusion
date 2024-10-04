# Code by Anbo Li, Fan Yang, Xinan Zhang, Xicai Pan, Xianli Xie
import os
import re
import cv2
import glob
import math
import shutil
import argparse
import numpy as np
import pandas as pd
from PIL import Image
import tensorflow as tf
import layoutparser as lp
from pdf2image import convert_from_path
from paddleocr import PaddleOCR, draw_ocr
from openpyxl import Workbook, load_workbook

# Recommended python version <= 3.8
# Before using, prepare GPU, download cuDNN and CUDA, for cuDNN and CUDA version compatibility refer to: https://developer.nvidia.com/rdp/cudnn-archive
# Paddle model installation: https://www.paddlepaddle.org.cn/
# Layoutparser installation: Download whl and then use pip install "E:\Edge\layoutparser-0.0.0-py3-none-any.whl"
# PaddleOCR installation: pip install "paddleocr>=2.0.1"

# Paddle installation check:
# import paddle
# paddle.utils.run_check()

# Program output: base folder\\Excel_union: merged Excel files
# Program output: base folder\\Data: stores intermediate steps
# Program output: base folder\\Data\\PDF name\\PDF name_results\\figure: cropped figures
# Program output: base folder\\Data\\PDF name\\PDF name_results\\table: cropped tables
# If the program crashes, try adding the following code:
# os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# !!! Folder path must not contain Chinese characters. If errors occur, try removing spaces, '-', and make filenames less than 10 letters long,etc.
# For parameter passing, please refer to the comment section at the end of the code.

# After restarting the interpreter, loading models is required. The program runtime is long, but generally does not exceed 5 minutes (referring to the time until prompt messages appear).

ocr = PaddleOCR(lang="en")

def ocr_and_crop(pdf_path, keywords, results_folder):
    """
    Extracts specific keywords from a PDF and crops the corresponding image regions.

    Parameters:
    pdf_path (str): Path to the PDF file.
    keywords (list of str): List of keywords to search for.
    results_folder (str): Folder path to store the cropped results.

    Returns:
    No return value. The cropped images will be saved to the specified results_folder.
    """
    def find_keyword_in_output(output, keyword):
        for item in output:
            text_block = item[1]
            text = text_block[0]
            if keyword.lower() in text.lower():
                return item
        return None

    def get_coords_of_keyword(output, keyword):
        item = find_keyword_in_output(output, keyword)
        if item:
            return item[0]
        return None

    # Copy and rename the PDF file
    folder_path = os.path.dirname(pdf_path)
    pdf_name_with_extension = os.path.basename(pdf_path)
    pdf_name_without_extension, extension = os.path.splitext(pdf_name_with_extension)

    # Use regular expression to match all numbers and letters
    new_pdf_base_name = ''.join(re.findall(r'[A-Za-z0-9]', pdf_name_without_extension))[:15]  # Slice to ensure no more than 15 characters

    # Add "_copy" suffix to the file name
    new_pdf_name = new_pdf_base_name + "_copy" + extension

    # Construct new PDF path
    new_pdf_path = os.path.join(folder_path, new_pdf_name)

    # Copy file to new path
    shutil.copyfile(pdf_path, new_pdf_path)

    images_path = os.path.join(folder_path, results_folder, 'images')
    clip_path = os.path.join(folder_path, results_folder, 'clip')
    os.makedirs(f"{images_path}", exist_ok=True)
    os.makedirs(f"{clip_path}", exist_ok=True)
    images = convert_from_path(new_pdf_path)

    # Remove '_copy.pdf' part from new_pdf_name
    image_copy_name = os.path.splitext(new_pdf_name)[0]

    # Then remove the ending '_copy'
    if image_copy_name.endswith('_copy'):
        image_copy_name = image_copy_name[:-5]

    for i, image in enumerate(images):
        image.save(f"{images_path}\\{image_copy_name}_page{i+1}.jpg", "JPEG")
        print(f"Saved:{image_copy_name}_page{i+1}.jpg")
        image_cv = cv2.imread(f"{images_path}\\{image_copy_name}_page{i+1}.jpg")
        output = ocr.ocr(f"{images_path}\\{image_copy_name}_page{i+1}.jpg")[0]

        coords_up, coords_down = None, None

        for keyword in keywords:
            if keyword == keywords[0]:
                for item in output:
                    text_block = item[1]
                    text = text_block[0]
                    if keyword.lower() in text.lower():
                        index = text.lower().find(keyword.lower())
                        text_before_keyword = text[:index].strip()
                        break
                else:
                    text_before_keyword = str(i+1)
            else:
                coords = get_coords_of_keyword(output, keyword)
                if coords:
                    if keyword == keywords[1]:
                        coords_up = coords[2][1]
                    elif keyword == keywords[2]:
                        coords_down = coords[0][1]

        if coords_up is None or coords_down is None:
            continue
        else:
            y1, y2 = int(min(coords_up, coords_down)), int(max(coords_up, coords_down))
            cropped_image = image_cv[y1:y2, :]
            # Use text_before_keyword as the filename
            cv2.imwrite(f"{clip_path}\\{text_before_keyword}.jpg", cropped_image)

    # Delete the copied PDF file
    os.remove(new_pdf_path)

def image_ocr(image_path, excel_folder):
    """
    Perform OCR on an image and save the results to an Excel file.

    Parameters:
    image_path (str): Path to the image file.
    excel_folder (str): Folder path to save the Excel file.

    Returns:
    No return value. The recognition results will be saved as an Excel file.
    """
    # Extract file name from image_path (without extension)
    number = os.path.splitext(os.path.basename(image_path))[0]
    image_cv = cv2.imread(image_path)
    image_height = image_cv.shape[0]
    image_width = image_cv.shape[1]
    output = ocr.ocr(image_path)[0]

    # Split the content of output into boxes, texts, probabilities
    boxes = [line[0] for line in output]
    texts = [line[1][0] for line in output]
    probabilities = [line[1][1] for line in output]

    image_boxes = image_cv.copy()
    for box, text in zip(boxes, texts):
        cv2.rectangle(
            image_boxes,
            (int(box[0][0]), int(box[0][1])),
            (int(box[2][0]), int(box[2][1])),
            (0, 0, 255),
            1,
        )
        cv2.putText(
            image_boxes,
            text,
            (int(box[0][0]), int(box[0][1])),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (255, 0, 0),
            1,
        )

    # --------------------Reconstruction--------------------
    # Get Horizontal and Vertical Lines
    im = image_cv.copy()
    horiz_boxes = []
    vert_boxes = []

    for box in boxes:
        x_h, x_v = 0, int(box[0][0])
        y_h, y_v = int(box[0][1]), 0
        width_h, width_v = image_width, int(box[2][0] - box[0][0])
        height_h, height_v = int(box[2][1] - box[0][1]), image_height

        horiz_boxes.append([x_h, y_h, x_h + width_h, y_h + height_h])
        vert_boxes.append([x_v, y_v, x_v + width_v, y_v + height_v])

        cv2.rectangle(
            im, (x_h, y_h), (x_h + width_h, y_h + height_h), (0, 0, 255), 1
        )
        cv2.rectangle(
            im, (x_v, y_v), (x_v + width_v, y_v + height_v), (0, 255, 0), 1
        )

    # Non-Max Suppression
    horiz_out = tf.image.non_max_suppression(
        horiz_boxes,
        probabilities,
        max_output_size=1000,
        iou_threshold=0.1,
        score_threshold=float("-inf"),
        name=None,
    )
    horiz_lines = np.sort(np.array(horiz_out))

    im_nms = image_cv.copy()
    for val in horiz_lines:
        cv2.rectangle(
            im_nms,
            (int(horiz_boxes[val][0]), int(horiz_boxes[val][1])),
            (int(horiz_boxes[val][2]), int(horiz_boxes[val][3])),
            (0, 0, 255),
            1,
        )

    vert_out = tf.image.non_max_suppression(
        vert_boxes,
        probabilities,
        max_output_size=1000,
        iou_threshold=0.05,
        score_threshold=float("-inf"),
        name=None,
    )
    vert_lines = np.sort(np.array(vert_out))

    for val in vert_lines:
        cv2.rectangle(
            im_nms,
            (int(vert_boxes[val][0]), int(vert_boxes[val][1])),
            (int(vert_boxes[val][2]), int(vert_boxes[val][3])),
            (255, 0, 0),
            1,
        )

    # Convert to EXCEL
    out_array = [
        ["" for i in range(len(vert_lines))] for j in range(len(horiz_lines))
    ]

    unordered_boxes = []

    for i in vert_lines:
        unordered_boxes.append(vert_boxes[i][0])
    ordered_boxes = np.argsort(unordered_boxes)

    def intersection(box_1, box_2):
        return [box_2[0], box_1[1], box_2[2], box_1[3]]

    def iou(box_1, box_2):
        x_1 = max(box_1[0], box_2[0])
        y_1 = max(box_1[1], box_2[1])
        x_2 = min(box_1[2], box_2[2])
        y_2 = min(box_1[3], box_2[3])

        inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
        if inter == 0:
            return 0

        box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
        box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

        return inter / float(box_1_area + box_2_area - inter)

    for i in range(len(horiz_lines)):
        for j in range(len(vert_lines)):
            resultant = intersection(
                horiz_boxes[horiz_lines[i]],
                vert_boxes[vert_lines[ordered_boxes[j]]],
            )

            for b in range(len(boxes)):
                the_box = [
                    boxes[b][0][0],
                    boxes[b][0][1],
                    boxes[b][2][0],
                    boxes[b][2][1],
                ]
                if iou(resultant, the_box) > 0.1:
                    out_array[i][j] = texts[b]

    out_array = np.array(out_array)
    df = pd.DataFrame(out_array)

    excel_file_path = os.path.join(excel_folder, f"{number}.xlsx")
    df.to_excel(excel_file_path, index=False)

    # Load Excel file
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    # Delete the first row
    sheet.delete_rows(1)

    # Save the modified file
    workbook.save(excel_file_path)

def process_images_in_folder(folder_path, excel_folder):
    """
    Perform OCR on all image files in a folder and save the results to Excel files.

    Parameters:
    folder_path (str): Folder path containing image files.
    excel_folder (str): Folder path to save the Excel files.

    Returns:
    No return value. An Excel file is generated for each image.
    """
    # Get all .jpg files in folder_path
    for image_path in glob.glob(os.path.join(folder_path, "*.jpg")):
        image_ocr(image_path, excel_folder)
        print(f"Processed and created Excel for {image_path}")

def merge_excels_to_sheets(source_folder, output_excel_path):
    """
    Merge all Excel files in a folder into a single Excel file, each original file as a separate sheet.

    Parameters:
    source_folder (str): Folder path containing source Excel files.
    output_excel_path (str): Path for the output Excel file.

    Returns:
    No return value, the merged Excel file is saved to output_excel_path.
    """
    # Create an Excel writer
    writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')

    # Iterate over all Excel files in source_folder
    for excel_file in glob.glob(os.path.join(source_folder, "*.xlsx")):
        # Use the original Excel file name (without extension) as the sheet name
        sheet_name = os.path.splitext(os.path.basename(excel_file))[0]

        # Read the Excel file into a DataFrame
        df = pd.read_excel(excel_file)

        # Write the DataFrame to a new sheet in the output Excel file
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the new Excel file
    writer.close()

def setup_environment(base_path):
    """
    Set up the environment for running the program, creating necessary folders.

    Parameters:
    base_path (str): Base path, all output folders will be created under this path.

    Returns:
    No return value, creates the necessary folder structure.
    """
    # Prevent system crash
    # os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    os.makedirs(os.path.join(base_path, "Excel_union"), exist_ok=True)
    os.makedirs(os.path.join(base_path, "Data"), exist_ok=True)

# Rename files to avoid errors
def rename_pdf_files(pdf_folder):
    """
    Rename PDF files, retaining letters and numbers, to avoid errors caused by file names.

    Parameters:
    pdf_folder (str): Folder path containing the PDF files.

    Returns:
    No return value, renaming is performed directly in the file system.
    """
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    for pdf_path in pdf_files:
        pdf_name_with_extension = os.path.basename(pdf_path)
        # Filter file name, keep only letters and numbers (spaces may cause reading errors)
        filtered_name = re.sub(r'[^a-zA-Z0-9]', '', os.path.splitext(pdf_name_with_extension)[0])
        filtered_name = re.sub(r'\s+', ' ', filtered_name)
        # Remove trailing spaces from the file name
        filtered_name = filtered_name.rstrip()

        # If the filtered file name exceeds 20 characters, truncate it to keep within 20 characters
        if len(filtered_name) > 20:
            filtered_name = filtered_name[:20]

        # Construct new PDF path
        filtered_pdf_path = os.path.join(pdf_folder, f'{filtered_name}.pdf')

        # Rename the PDF file
        if pdf_path != filtered_pdf_path:
            shutil.move(pdf_path, filtered_pdf_path)

def copy_pdf_folder(original_folder, backup_folder):
    """
    Copy an entire PDF folder to a backup location.

    Parameters:
    original_folder (str): Path to the source PDF folder.
    backup_folder (str): Path to the backup folder.

    Returns:
    No return value, the folder copy operation is performed directly in the file system.
    """
    if not os.path.exists(backup_folder):
        shutil.copytree(original_folder, backup_folder)

def restore_original_pdf_folder(backup_folder, original_folder):
    """
    Restore the original PDF folder from the backup folder.

    Parameters:
    backup_folder (str): Path to the backup folder.
    original_folder (str): Path to the original folder, which will be restored.

    Returns:
    No return value, restores the backup folder to the original folder.
    """
    if os.path.exists(original_folder):
        shutil.rmtree(original_folder)
    shutil.move(backup_folder, original_folder)

def intersection(box_1, box_2):
    """
    Calculate the intersection of two bounding boxes.

    Parameters:
    box_1 (list of int): Coordinates of the first bounding box [x1, y1, x2, y2].
    box_2 (list of int): Coordinates of the second bounding box [x1, y1, x2, y2].

    Returns:
    list: Coordinates of the intersected region [x1, y1, x2, y2].
    """
    return [box_2[0], box_1[1], box_2[2], box_1[3]]

# Calculate IoU (Intersection over Union) for OCR
def iou(box_1, box_2):
    """
    Calculate the Intersection over Union (IoU) of two bounding boxes.

    Parameters:
    box_1 (list of int): Coordinates of the first bounding box [x1, y1, x2, y2].
    box_2 (list of int): Coordinates of the second bounding box [x1, y1, x2, y2].

    Returns:
    float: IoU value of the two bounding boxes, ranging from 0 to 1.
    """
    x_1 = max(box_1[0], box_2[0])
    y_1 = max(box_1[1], box_2[1])
    x_2 = min(box_1[2], box_2[2])
    y_2 = min(box_1[3], box_2[3])

    inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
    if inter == 0:
        return 0

    box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
    box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

    return inter / float(box_1_area + box_2_area - inter)

def copy_sheet(source_sheet, target_sheet):
    """
    Copy the content from the source worksheet to the target worksheet.

    Parameters:
    source_sheet (openpyxl.worksheet.worksheet.Worksheet): Source worksheet object.
    target_sheet (openpyxl.worksheet.worksheet.Worksheet): Target worksheet object.

    Returns:
    No return value, content is copied to the target worksheet.
    """
    for row in source_sheet:
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

# Create corresponding folder for each PDF
def img_initialization(pages_dir, results_dir, images, pdf_name_without_extension, table_ok, figure_ok):
    """
    Initialize the directory structure for image processing and save the images of each page of the PDF.

    Parameters:
    pages_dir (str): Folder path to save processed images.
    results_dir (str): Folder path to save result files.
    images (list of PIL.Image): List of images converted from PDF.
    pdf_name_without_extension (str): PDF file name (without extension).
    table_ok (bool): Whether to extract table information.
    figure_ok (bool): Whether to extract figure information.

    Returns:
    No return value, initializes folder structure and saves images.
    """
    os.makedirs(pages_dir, exist_ok=True)
    os.makedirs(results_dir, exist_ok=True)
    if figure_ok:
        os.makedirs(f"{results_dir}\\figure", exist_ok=True)
    if table_ok:
        os.makedirs(f"{results_dir}\\table", exist_ok=True)
        os.makedirs(f"{results_dir}\\detections", exist_ok=True)
        os.makedirs(f"{results_dir}\\horiz_vert", exist_ok=True)
        os.makedirs(f"{results_dir}\\im_nms", exist_ok=True)
        os.makedirs(f"{results_dir}\\excel", exist_ok=True)

    for i in range(len(images)):
        images[i].save(
            f"{pages_dir}\\{pdf_name_without_extension}_page" + str(i) + ".jpg", "JPEG"
        )
        print(f"{pages_dir}\\{pdf_name_without_extension}_page" + str(i) + ".jpg")

# Paddle model parses PDF images, classified as 0: "Text", 1: "Title", 2: "List", 3: "Table", 4: "Figure"
def pdf_detection(pages_dir, pdf_name_without_extension, page_num, model_threshold):
    """
    Perform object detection on PDF pages, extracting text, table, and figure regions.

    Parameters:
    pages_dir (str): Folder path to save page images.
    pdf_name_without_extension (str): PDF file name (without extension).
    page_num (int): The current page number being processed.
    model_threshold (float): Confidence threshold for model detection.

    Returns:
    layout (list): Detected object layout information, including text, table, and figure.
    """
    image_path = f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg"
    print(image_path)

    image = cv2.imread(image_path)
    image = image[..., ::-1]
    print(f"Processing file: {pdf_name_without_extension}, page: {page_num}")

    # Load model
    model = lp.PaddleDetectionLayoutModel(
        config_path="lp://PubLayNet/ppyolov2_r50vd_dcn_365e_publaynet/config",
        # Confidence of 0.5 or higher is considered as a specific class
        threshold = model_threshold,
        label_map={0: "Text", 1: "Title", 2: "List", 3: "Table", 4: "Figure"},
        enforce_cpu=False,
        enable_mkldnn=True,
    )  # math kernel library
    # Detect
    layout = model.detect(image)
    return layout

# Merge all generated xlsx files for a single PDF
def create_excel(results_dir, pdf_name_without_extension, base_path):
    """
    Merge all generated Excel files for a single PDF into one.

    Parameters:
    results_dir (str): Directory storing intermediate Excel files.
    pdf_name_without_extension (str): PDF file name (without extension).
    base_path (str): Base path to save the final merged Excel file.

    Returns:
    No return value, the merged Excel file is saved to base_path.
    """
    wb = Workbook()
    excel_dir = f"{results_dir}\\excel"
    final_excel_path = os.path.join(base_path, f"Excel_union\\{pdf_name_without_extension}.xlsx")

    excel_files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx")]

    # If there are no Excel files in the directory, add a default sheet
    if not excel_files:
        wb.create_sheet(title="Sheet1")
    else:
        # If there are Excel files, remove the default added sheet
        wb.remove(wb.active)

    for excel_file in excel_files:
        file_path = os.path.join(excel_dir, excel_file)
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active

        # Use file name (excluding extension) as the sheet name
        sheet_name = "_".join(excel_file.split("_")[1:])
        sheet_name = os.path.splitext(sheet_name)[0]

        # Limit sheet name to the first 25 characters
        actual_sheet_name = sheet_name[:20]

        # Create a new sheet with the actual sheet name
        target_sheet = wb.create_sheet(title=actual_sheet_name)

        # Copy content to the new sheet
        copy_sheet(source_sheet, target_sheet)

        # Find the last row in the target sheet
        last_row = target_sheet.max_row

        # Insert the file name (excluding extension) in the second row after all copied content
        target_sheet.cell(row=last_row + 2, column=1, value=sheet_name)

    # Save the final workbook
    wb.save(final_excel_path)

def process_pdf_file(pdf_path, base_path, table_ok, figure_ok, model_threshold):
    """
    Process a PDF file, extract tables and figures, and generate corresponding Excel files.

    Parameters:
    pdf_path (str): Path to the PDF file.
    base_path (str): Base path for saving the processing results.
    table_ok (bool): Whether to extract table information.
    figure_ok (bool): Whether to extract figure information.
    model_threshold (float): Confidence threshold for model detection.

    Returns:
    No return value, processing results are saved to the specified path.
    """
    pdf_name_with_extension = os.path.basename(pdf_path)
    pdf_name_without_extension = os.path.splitext(pdf_name_with_extension)[0]

    images = convert_from_path(pdf_path)

    print("image finish")

    pages_dir = os.path.join(base_path, f"Data\\{pdf_name_without_extension}\\{pdf_name_without_extension}_pages")
    results_dir = (
        os.path.join(base_path, f"Data\\{pdf_name_without_extension}\\{pdf_name_without_extension}_results")
    )

    img_initialization(pages_dir, results_dir, images, pdf_name_without_extension, table_ok, figure_ok)

    # --------------------Table Extraction--------------------
    # Iterate over each page, locate tables and figures
    for page_idx in range(len(images)):
        page_num = f"page{page_idx}"

        layout = pdf_detection(pages_dir, pdf_name_without_extension, page_num, model_threshold)

        table_loc = []
        figure_loc = []

        # Get table bounding box coordinates (slightly expanded)
        for l in layout:
            # Table location
            if table_ok:
                if l.type == "Table":
                    x_1 = int(l.block.x_1) - 4
                    y_1 = math.ceil(l.block.y_1) + 4
                    x_2 = math.ceil(l.block.x_2) + 4
                    y_2 = int(l.block.y_2) - 4
                    table_loc.append((x_1, y_1, x_2, y_2))
            # Figure location
            if figure_ok:
                if l.type == 'Figure':
                    x_3 = int(l.block.x_1) - 4
                    y_3 = math.ceil(l.block.y_1) + 4
                    x_4 = math.ceil(l.block.x_2) + 4
                    y_4 = int(l.block.y_2) - 4
                    figure_loc.append((x_3, y_3, x_4, y_4))

        if figure_ok:
            for idx, cor in enumerate(figure_loc):
                x_3, y_3, x_4, y_4 = cor

                im = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\figure\\{pdf_name_without_extension}_{page_num}_figure_{idx}.jpg",
                    im[y_3:y_4, x_3:x_4],
                )
        if table_ok:
            for idx, cor in enumerate(table_loc):
                x_1, y_1, x_2, y_2 = cor
                combined_text = None
                im_text = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_text_{idx}.jpg",
                    im_text[0:int(y_1), int(x_1):int(x_2)],
                )
                ocr = PaddleOCR(lang="en")
                image_path = f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_text_{idx}.jpg"
                image_cv = cv2.imread(image_path)
                image_height = image_cv.shape[0]
                image_width = image_cv.shape[1]
                output_table = ocr.ocr(image_path)[0]
                if output_table and output_table[0]:  # Ensure output_table is not None and contains results
                    texts_table = [line_table[1][0] for line_table in output_table]
                    texts_table_lower = [text.lower() for text in texts_table]

                    # Initialize a variable to hold the index of 'table' if found
                    table_index = None

                    # Iterate over texts_table_lower in reverse to find 'table'
                    for i, text in enumerate(reversed(texts_table_lower)):
                        if 'table' in text:
                            # Calculate the correct index in the original list
                            table_index = len(texts_table_lower) - 1 - i
                            break

                    # Check if 'table' was found
                    if table_index is not None:
                        # Extract and print the texts from 'table' onwards
                        result_texts = texts_table[table_index:]
                        combined_text = ' '.join(result_texts)
                        combined_text = combined_text[:140]
                    else:
                        combined_text = None

                    if combined_text is not None:
                        # Replace all characters not a letter, a number, or a space with '_'
                        combined_text = re.sub(r'[^a-zA-Z0-9\s]', '_', combined_text)

                im = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_table_{idx}.jpg",
                    im[y_1:y_2, x_1:x_2],
                )

                # --------------------Text Detection and Recognition--------------------
                # OCR part, can be independent
                ocr = PaddleOCR(lang="en")
                image_path = f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_table_{idx}.jpg"
                image_cv = cv2.imread(image_path)
                image_height = image_cv.shape[0]
                image_width = image_cv.shape[1]
                output = ocr.ocr(image_path)[0]

                # Split the content of output into boxes, texts, probabilities
                boxes = [line[0] for line in output]
                texts = [line[1][0] for line in output]
                probabilities = [line[1][1] for line in output]

                image_boxes = image_cv.copy()
                for box, text in zip(boxes, texts):
                    cv2.rectangle(
                        image_boxes,
                        (int(box[0][0]), int(box[0][1])),
                        (int(box[2][0]), int(box[2][1])),
                        (0, 0, 255),
                        1,
                    )
                    cv2.putText(
                        image_boxes,
                        text,
                        (int(box[0][0]), int(box[0][1])),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.6,
                        (255, 0, 0),
                        1,
                    )
                cv2.imwrite(
                    f"{results_dir}\\detections\\{pdf_name_without_extension}_{page_num}_detections_{idx}.jpg",
                    image_boxes,
                )

                # --------------------Reconstruction--------------------
                # Get Horizontal and Vertical Lines
                im = image_cv.copy()
                horiz_boxes = []
                vert_boxes = []

                for box in boxes:
                    x_h, x_v = 0, int(box[0][0])
                    y_h, y_v = int(box[0][1]), 0
                    width_h, width_v = image_width, int(box[2][0] - box[0][0])
                    height_h, height_v = int(box[2][1] - box[0][1]), image_height

                    horiz_boxes.append([x_h, y_h, x_h + width_h, y_h + height_h])
                    vert_boxes.append([x_v, y_v, x_v + width_v, y_v + height_v])

                    cv2.rectangle(
                        im, (x_h, y_h), (x_h + width_h, y_h + height_h), (0, 0, 255), 1
                    )
                    cv2.rectangle(
                        im, (x_v, y_v), (x_v + width_v, y_v + height_v), (0, 255, 0), 1
                    )
                cv2.imwrite(
                    f"{results_dir}\\horiz_vert\\{pdf_name_without_extension}_{page_num}_horiz_vert_{idx}.jpg",
                    im,
                )

                # Non-Max Suppression
                horiz_out = tf.image.non_max_suppression(
                    horiz_boxes,
                    probabilities,
                    max_output_size=1000,
                    iou_threshold=0.1,
                    score_threshold=float("-inf"),
                    name=None,
                )
                horiz_lines = np.sort(np.array(horiz_out))

                im_nms = image_cv.copy()
                for val in horiz_lines:
                    cv2.rectangle(
                        im_nms,
                        (int(horiz_boxes[val][0]), int(horiz_boxes[val][1])),
                        (int(horiz_boxes[val][2]), int(horiz_boxes[val][3])),
                        (0, 0, 255),
                        1,
                    )

                vert_out = tf.image.non_max_suppression(
                    vert_boxes,
                    probabilities,
                    max_output_size=1000,
                    iou_threshold=0.05,
                    score_threshold=float("-inf"),
                    name=None,
                )
                vert_lines = np.sort(np.array(vert_out))

                for val in vert_lines:
                    cv2.rectangle(
                        im_nms,
                        (int(vert_boxes[val][0]), int(vert_boxes[val][1])),
                        (int(vert_boxes[val][2]), int(vert_boxes[val][3])),
                        (255, 0, 0),
                        1,
                    )
                cv2.imwrite(
                    f"{results_dir}\\im_nms\\{pdf_name_without_extension}_{page_num}_im_nms_{idx}.jpg",
                    im_nms,
                )

                # Convert to EXCEL
                out_array = [
                    ["" for i in range(len(vert_lines))] for j in range(len(horiz_lines))
                ]

                unordered_boxes = []

                for i in vert_lines:
                    unordered_boxes.append(vert_boxes[i][0])
                ordered_boxes = np.argsort(unordered_boxes)

                def intersection(box_1, box_2):
                    return [box_2[0], box_1[1], box_2[2], box_1[3]]

                def iou(box_1, box_2):
                    x_1 = max(box_1[0], box_2[0])
                    y_1 = max(box_1[1], box_2[1])
                    x_2 = min(box_1[2], box_2[2])
                    y_2 = min(box_1[3], box_2[3])

                    inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
                    if inter == 0:
                        return 0

                    box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
                    box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

                    return inter / float(box_1_area + box_2_area - inter)

                for i in range(len(horiz_lines)):
                    for j in range(len(vert_lines)):
                        resultant = intersection(
                            horiz_boxes[horiz_lines[i]],
                            vert_boxes[vert_lines[ordered_boxes[j]]],
                        )

                        for b in range(len(boxes)):
                            the_box = [
                                boxes[b][0][0],
                                boxes[b][0][1],
                                boxes[b][2][0],
                                boxes[b][2][1],
                            ]
                            if iou(resultant, the_box) > 0.1:
                                out_array[i][j] = texts[b]

                out_array = np.array(out_array)
                df = pd.DataFrame(out_array)
                df.to_excel(
                    f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx",
                    index=False,
                )

                # Load the Excel file
                workbook = load_workbook(
                    filename=f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx"
                )
                sheet = workbook.active

                # Delete the first row
                sheet.delete_rows(1)

                # Save the modified file
                workbook.save(
                    f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx"
                )
    if table_ok:
        create_excel(results_dir, pdf_name_without_extension, base_path)

def main(base_path, table_ok, figure_ok, model_threshold, mode, base_folder, keywords):
    """
    Main function to control the workflow of the program, processing PDF files to extract tables or keyword-related image areas.

    Parameters:
    base_path (str): Base path to store processing results.
    table_ok (bool): Whether to extract table information.
    figure_ok (bool): Whether to extract figure information.
    model_threshold (float): Confidence threshold for model detection.
    mode (str): Processing mode, either 'normal' or 'keyword'.
    base_folder (str): Base folder path where PDF files are stored.
    keywords (list of str): List of keywords to search for (only valid in 'keyword' mode).

    Returns:
    No return value, completes PDF file processing in the specified mode.
    """
    print(mode)
    if mode == 'normal':
        backup_folder = os.path.join(base_path, "Document_Copy")

        setup_environment(base_path)

        # Create Document folder (clear if it already exists)
        pdf_folder = os.path.join(base_path, "Document")
        if os.path.exists(pdf_folder):
            shutil.rmtree(pdf_folder)
        os.makedirs(pdf_folder)

        # Copy PDF files from base_path to Document folder
        for pdf_file in glob.glob(os.path.join(base_path, "*.pdf")):
            shutil.copy(pdf_file, pdf_folder)

        # Copy the PDF folder
        copy_pdf_folder(pdf_folder, backup_folder)

        rename_pdf_files(pdf_folder)

        # Get all PDF files in the folder
        pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))

        # Iterate through PDF files for cropping and OCR
        for pdf_path in pdf_files:
            process_pdf_file(pdf_path, base_path, table_ok, figure_ok, model_threshold)

        restore_original_pdf_folder(backup_folder, pdf_folder)

        # Delete the Document folder
        shutil.rmtree(pdf_folder)

    if mode == 'keyword':
        for pdf_file in os.listdir(base_folder):
            if pdf_file.endswith('.pdf'):
                pdf_path = os.path.join(base_folder, pdf_file)

                pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
                results_folder = f'{pdf_name}_results'
                ocr_and_crop(pdf_path, keywords, results_folder)

                base_path = os.path.join(base_folder, results_folder)
                clip_folder_path = os.path.join(base_path, 'clip')
                excel_path = os.path.join(base_path, "union.xlsx")
                excel_folder = os.path.join(base_path, "excel")
                if not os.path.exists(excel_folder):
                    os.makedirs(excel_folder)

                if os.listdir(clip_folder_path):
                    # If the clip folder is not empty, proceed with the following functions
                    process_images_in_folder(clip_folder_path, excel_folder)
                    merge_excels_to_sheets(excel_folder, excel_path)

if __name__ == "__main__":
    # Script argument parsing, or adjust the default parameters for running
    parser = argparse.ArgumentParser()
    # Path where PDF files are stored, recommended to use '\\'
    # parser.add_argument("--base_path", type=str, default=os.path.join(os.getcwd(), "test"))
    parser.add_argument("--base_path", type=str, default=r"F:\\OCR\\test")
    # Whether to extract tables
    parser.add_argument("--table_ok", type=bool, default=True)
    # Whether to extract figures
    parser.add_argument("--figure_ok", type=bool, default=True)
    # Confidence threshold for model detection
    parser.add_argument("--model_threshold", type=float, default=0.5)
    # Processing mode ('normal' is used for general papers, the keyword function is not fully developed)
    parser.add_argument("--mode", type=str, default='normal')
    # Path where files are stored in keyword mode
    parser.add_argument('--base_folder', type=str, default=r"base_folder")
    # List of keywords
    parser.add_argument('--keywords', type=str, nargs='+', default=["Keywords"])

    args = parser.parse_args()

    main(args.base_path, args.table_ok, args.figure_ok, args.model_threshold, args.mode, args.base_folder, args.keywords)
