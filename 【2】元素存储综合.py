import pandas as pd
import re
import os
import csv
import openpyxl
import easygui
import glob


def remove_sub_superscripts(s):
    """
    删除字符串中的上下标字符和特殊符号。

    参数:
    s (str): 需要处理的字符串。

    返回:
    str: 删除上下标和特殊符号后的字符串。
    """
    s = re.sub(r'[\u2070-\u209F\u00B2\u00B3\u00B9\u00BC-\u00BE]+', '', s)
    s = re.sub(r'[\u2080-\u208E]+', '', s)
    s = s.replace(' ', '').replace('-', '').replace('+', '').replace('"', '').replace("'", '')
    return s


def get_elements_from_user():
    """
    弹出对话框让用户输入关键词和文件夹路径。

    返回:
    tuple: 包含关键词列表、待处理文件夹路径和CSV存储路径的元组。
    """
    message = "请输入希望存储的特殊关键词，用空格隔开。然后输入文件夹路径。"
    title = "数据存储器"
    # 关键词即无单位会干扰识别的元素
    default_keywords = "ph cn tc fc temp depth eh"
    default_dir_path = r"C:\\"
    default_dir_path1 = r"C:\\"

    fields = [
        "关键词（用空格隔开）",
        "待提取表格所在的文件夹路径",
        "想要放置生成CSV的路径"
    ]

    default_values = [
        default_keywords,
        default_dir_path,
        default_dir_path1
    ]

    # 使用 easygui 弹出多输入框获取用户输入
    user_input = easygui.multenterbox(message, title, fields, default_values)
    if user_input is not None:
        keywords = user_input[0].split()
        dir_path = user_input[1].replace("\\", "\\\\")
        dir_path1 = user_input[2].replace("\\", "\\\\")
        return keywords, dir_path, dir_path1
    else:
        return [], "", ""


# 获取用户输入的关键词列表、文件夹路径及CSV路径
keyword_list, dir_path, dir_path1 = get_elements_from_user()


def rename_files_in_folder(folder_path):
    """
    重命名指定文件夹中的所有 .xlsx 文件，限制文件名长度并去除不必要的后缀。

    参数:
    folder_path (str): 包含 .xlsx 文件的文件夹路径。
    """
    # 获取所有 .xlsx 文件
    xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

    for file_path in xlsx_files:
        # 获取文件的目录和文件名
        directory, filename_with_ext = os.path.split(file_path)
        filename, extension = os.path.splitext(filename_with_ext)

        # 删除文件名中的 "converted" 后缀
        if filename.endswith('converted'):
            filename = filename[:-len('converted')]

        # 限制文件名长度为40个字符以内
        if len(filename) > 40:
            filename = filename[:40]

        # 构造新的文件路径
        new_file_path = os.path.join(directory, filename + extension)

        # 重命名文件（注释掉实际重命名代码）
        # if new_file_path != file_path:
        #     os.rename(file_path, new_file_path)


# 使用上述函数重命名文件
rename_files_in_folder(dir_path)


def transpose(df):
    """
    转置给定的 DataFrame。

    参数:
    df (pd.DataFrame): 需要转置的 DataFrame。

    返回:
    pd.DataFrame: 转置后的 DataFrame。
    """
    df_transposed = df.transpose()
    return df_transposed


# 遍历文件夹中的所有 .xlsx 文件
for filename in os.listdir(dir_path):
    if not filename.endswith('.xlsx'):
        continue
    file_path = os.path.join(dir_path, filename)

    # 加载工作簿
    book = openpyxl.load_workbook(file_path)

    sheets_data = {}
    # 遍历每个工作表
    for sheet in book.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        header_row = None
        # 检测表头行
        for i, row in df.iterrows():
            metal_count = sum([1 for cell in row if re.search(r'\(.*(/|ppm|ppb).*\)', str(cell)) or any(
                keyword in str(cell).lower() for keyword in keyword_list)])
            if metal_count >= 2:
                header_row = i
                break

        # 如果没有检测到表头，则转置工作表数据
        if header_row is None:
            df = transpose(df)
            sheets_data[sheet + ""] = df  # 将转置后的数据存入字典
        else:
            sheets_data[sheet] = df  # 存储原始数据

    # 将处理后的数据写回到 Excel 文件中
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet, df in sheets_data.items():
            df.to_excel(writer, sheet_name=sheet, index=False, header=False)

# 定义 CSV 文件的列名
csv_columns = ['filename', 'element', 'sample_info', 'unit', 'data']
# 创建并写入 CSV 文件的头部
with open(os.path.join(dir_path1, 'tuples_sheet.csv'), 'w', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(csv_columns)

# 遍历文件夹中的所有 .xlsx 文件
for filename in os.listdir(dir_path):
    if not filename.endswith('.xlsx'):
        continue
    file_path = os.path.join(dir_path, filename)
    book = openpyxl.load_workbook(file_path)

    # 打开输出文本文件
    with open(os.path.join(dir_path1, 'output.txt'), 'a') as f:
        for sheet in book.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None)
            sheet_text = df.to_string().lower()

            # 检测是否存在描述性数据（如均值、最大值等）
            if any(keyword in sheet_text for keyword in
                   ['mean', 'average', 'maximum', 'minimum', 'Maximum', 'Minimum']):
                output_str = f"‘{filename}-{sheet}’中存在描述性数据\n"
                print(output_str, end='')
                continue

            header_row = None
            # 检测表头行
            for i, row in df.iterrows():
                metal_count = sum([1 for cell in row if re.search(r'\(.*(/|ppm|ppb).*\)', str(cell)) or any(
                    keyword in str(cell).lower() for keyword in keyword_list)])
                if metal_count >= 2:
                    header_row = i
                    break

            # 处理检测到表头的工作表数据
            if header_row is not None:
                df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)
                cols_with_unit = [col for col in df.columns if re.search(r'\(.*(/|ppm|ppb).*\)', str(col)) or any(
                    keyword in str(col).lower() for keyword in keyword_list)]
                rearranged_data = []
                for i, row in df.iterrows():
                    sample_info = str(i + 1) + '(' + ' '.join(row.drop(cols_with_unit).astype(str)) + ')'
                    for col in cols_with_unit:
                        for keyword in keyword_list:
                            if keyword in col.lower():
                                metal = keyword.upper()
                                unit = None
                                data = row[col]
                                break
                        else:
                            metal = re.search(r'(.*)\(', col).group(1)
                            metal = metal.replace(' ', '')
                            unit = re.search(r'\((.*)\)', col).group(1)
                            unit = unit.replace(' ', '')
                            data = row[col]

                        if metal != "":
                            rearranged_row = [filename, metal, sample_info, unit, data]
                            rearranged_data.append(rearranged_row)

                # 将处理后的数据写入 CSV 文件
                with open(os.path.join(dir_path1, 'tuples_sheet.csv'), 'a', newline='',
                          encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerows(rearranged_data)
            else:
                # 如果没有找到符合条件的行或列，输出跳过信息
                output_str = f"‘{filename}-{sheet}’中无法找到符合条件的行或列，跳过处理\n"
                print(output_str, end='')
                f.write(output_str)
