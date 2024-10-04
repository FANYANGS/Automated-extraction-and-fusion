# Code by Anbo Li, Fan Yang, Xinan Zhang, Xicai Pan, Xianli Xie
import pandas as pd
import re
import os
import csv
import openpyxl
import easygui
import glob

def remove_sub_superscripts(s):
    """
    Remove subscript, superscript characters, and special symbols from a string.

    Parameters:
    s (str): The string to be processed.

    Returns:
    str: The string after removing subscripts, superscripts, and special symbols.
    """
    s = re.sub(r'[\u2070-\u209F\u00B2\u00B3\u00B9\u00BC-\u00BE]+', '', s)
    s = re.sub(r'[\u2080-\u208E]+', '', s)
    s = s.replace(' ', '').replace('-', '').replace('+', '').replace('"', '').replace("'", '')
    return s

def get_elements_from_user():
    """
    Display a dialog box for the user to input keywords and folder paths.

    Returns:
    tuple: A tuple containing the list of keywords, the folder path to be processed, and the CSV storage path.
    """
    message = "Please enter the special keywords you wish to store, separated by spaces. Then enter the folder path."
    title = "Data Storage"
    # Keywords are elements without units that interfere with recognition
    default_keywords = "ph cn tc fc temp depth eh"
    default_dir_path = r"C:\\"
    default_dir_path1 = r"C:\\"

    fields = [
        "Keywords (separated by spaces)",
        "Folder path containing tables to extract",
        "Path to place the generated CSV"
    ]

    default_values = [
        default_keywords,
        default_dir_path,
        default_dir_path1
    ]

    # Use easygui to pop up a multi-input box to get user input
    user_input = easygui.multenterbox(message, title, fields, default_values)
    if user_input is not None:
        keywords = user_input[0].split()
        dir_path = user_input[1].replace("\\", "\\\\")
        dir_path1 = user_input[2].replace("\\", "\\\\")
        return keywords, dir_path, dir_path1
    else:
        return [], "", ""

# Get the list of keywords, folder path, and CSV path from user input
keyword_list, dir_path, dir_path1 = get_elements_from_user()

def rename_files_in_folder(folder_path):
    """
    Rename all .xlsx files in the specified folder, limit the file name length, and remove unnecessary suffixes.

    Parameters:
    folder_path (str): The folder path containing .xlsx files.
    """
    # Get all .xlsx files
    xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

    for file_path in xlsx_files:
        # Get the directory and file name of the file
        directory, filename_with_ext = os.path.split(file_path)
        filename, extension = os.path.splitext(filename_with_ext)

        # Remove the "converted" suffix from the file name
        if filename.endswith('converted'):
            filename = filename[:-len('converted')]

        # Limit the file name length to within 40 characters
        if len(filename) > 40:
            filename = filename[:40]

        # Construct a new file path
        new_file_path = os.path.join(directory, filename + extension)

        # Rename the file (commented out actual renaming code)
        # if new_file_path != file_path:
        #     os.rename(file_path, new_file_path)

# Use the above function to rename files
rename_files_in_folder(dir_path)

def transpose(df):
    """
    Transpose the given DataFrame.

    Parameters:
    df (pd.DataFrame): The DataFrame to be transposed.

    Returns:
    pd.DataFrame: The transposed DataFrame.
    """
    df_transposed = df.transpose()
    return df_transposed

# Iterate over all .xlsx files in the folder
for filename in os.listdir(dir_path):
    if not filename.endswith('.xlsx'):
        continue
    file_path = os.path.join(dir_path, filename)

    # Load the workbook
    book = openpyxl.load_workbook(file_path)

    sheets_data = {}
    # Iterate over each sheet
    for sheet in book.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        header_row = None
        # Detect the header row
        for i, row in df.iterrows():
            metal_count = sum([1 for cell in row if re.search(r'\(.*(/|ppm|ppb).*\)', str(cell)) or any(
                keyword in str(cell).lower() for keyword in keyword_list)])
            if metal_count >= 2:
                header_row = i
                break

        # If no header is detected, transpose the sheet data
        if header_row is None:
            df = transpose(df)
            sheets_data[sheet + ""] = df  # Store the transposed data in the dictionary
        else:
            sheets_data[sheet] = df  # Store the original data

    # Write the processed data back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet, df in sheets_data.items():
            df.to_excel(writer, sheet_name=sheet, index=False, header=False)

# Define the column names for the CSV file
csv_columns = ['filename', 'element', 'sample_info', 'unit', 'data']
# Create and write the header of the CSV file
with open(os.path.join(dir_path1, 'tuples_sheet.csv'), 'w', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(csv_columns)

# Iterate over all .xlsx files in the folder
for filename in os.listdir(dir_path):
    if not filename.endswith('.xlsx'):
        continue
    file_path = os.path.join(dir_path, filename)
    book = openpyxl.load_workbook(file_path)

    # Open the output text file
    with open(os.path.join(dir_path1, 'output.txt'), 'a') as f:
        for sheet in book.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None)
            sheet_text = df.to_string().lower()

            # Detect if there is descriptive data (such as mean, maximum, etc.)
            if any(keyword in sheet_text for keyword in
                   ['mean', 'average', 'maximum', 'minimum', 'Maximum', 'Minimum']):
                output_str = f"Descriptive data found in '{filename}-{sheet}'\n"
                print(output_str, end='')
                continue

            header_row = None
            # Detect the header row
            for i, row in df.iterrows():
                metal_count = sum([1 for cell in row if re.search(r'\(.*(/|ppm|ppb).*\)', str(cell)) or any(
                    keyword in str(cell).lower() for keyword in keyword_list)])
                if metal_count >= 2:
                    header_row = i
                    break

            # Process the sheet data with detected header
            if header_row is not None:
                df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)
                cols_with_unit = [col for col in df.columns if re.search(r'\(.*(/|ppm|ppb).*\)', str(col)) or any(
                    keyword in str(col).lower() for keyword in keyword_list)]
                rearranged_data = []
                for i, row in df.iterrows():
                    sample_info = str(i + 1) + '(' + ' '.join(row.drop(cols_with_unit).astype(str)) + ')'
                    for col in cols_with_unit:
                        for keyword in keyword_list:
                            if keyword in col.lower():
                                metal = keyword.upper()
                                unit = None
                                data = row[col]
                                break
                        else:
                            metal = re.search(r'(.*)\(', col).group(1)
                            metal = metal.replace(' ', '')
                            unit = re.search(r'\((.*)\)', col).group(1)
                            unit = unit.replace(' ', '')
                            data = row[col]

                        if metal != "":
                            rearranged_row = [filename, metal, sample_info, unit, data]
                            rearranged_data.append(rearranged_row)

                # Write the processed data to the CSV file
                with open(os.path.join(dir_path1, 'tuples_sheet.csv'), 'a', newline='',
                          encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerows(rearranged_data)
            else:
                # If no qualifying rows or columns are found, output skip information
                output_str = f"No qualifying rows or columns found in '{filename}-{sheet}', skipping processing\n"
                print(output_str, end='')
                f.write(output_str)
