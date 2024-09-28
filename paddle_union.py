import os
import re
import cv2
import glob
import math
import shutil
import argparse
import numpy as np
import pandas as pd
from PIL import Image
import tensorflow as tf
import layoutparser as lp
from pdf2image import convert_from_path
from paddleocr import PaddleOCR, draw_ocr
from openpyxl import Workbook, load_workbook

# 建议python=3.8及以下
# 使用前需准备好GPU,下载cuDNN与CUDA,cuDNN与CUDA版本联系参考:https:\\developer.nvidia.com\\rdp\\cudnn-archive
# Paddle模型安装:https:\\www.paddlepaddle.org.cn\\
# layoutparser安装:下载whl后 pip install "E:\Edge\layoutparser-0.0.0-py3-none-any.whl"
# paddleocr安装:pip install "paddleocr>=2.0.1"

# paddle安装检查:
# import paddle
# paddle.utils.run_check()

# 程序输出:base文件夹\\Excel_union: 合并后的Excel
# 程序输出:base文件夹\\Data: 存放各中间步骤
# 程序输出:base文件夹\\Data\\PDF名称\\PDF名称_results\\figure: 裁出的图
# 程序输出:base文件夹\\Data\\PDF名称\\PDF名称_results\\table: 裁出的表
# 若程序崩溃可尝试加如下代码：
# os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# 文件夹路径不能出现中文，若出现报错，可尝试去除空格等
# 关于参数传入请跳转到代码最后的注释部分

# 重启解释器后需要加载模型等，程序运行时间较长，但一般不会超过5分钟（指代码运行至出现提示信息）。


ocr = PaddleOCR(lang="en")

def ocr_and_crop(pdf_path, keywords, results_folder):
    """
    从PDF中提取指定关键词并裁剪出相应区域的图像。

    参数:
    pdf_path (str): PDF文件的路径。
    keywords (list of str): 需要查找的关键词列表。
    results_folder (str): 存储裁剪结果的文件夹路径。

    返回:
    无返回值，裁剪后的图像将保存到指定的results_folder中。
    """
    def find_keyword_in_output(output, keyword):
        for item in output:
            text_block = item[1]
            text = text_block[0]
            if keyword.lower() in text.lower():
                return item
        return None

    def get_coords_of_keyword(output, keyword):
        item = find_keyword_in_output(output, keyword)
        if item:
            return item[0]
        return None

    # 复制并重命名 PDF文件
    folder_path = os.path.dirname(pdf_path)
    pdf_name_with_extension = os.path.basename(pdf_path)
    pdf_name_without_extension, extension = os.path.splitext(pdf_name_with_extension)

    # 使用正则表达式匹配所有数字和字母
    new_pdf_base_name = ''.join(re.findall(r'[A-Za-z0-9]', pdf_name_without_extension))[:15]  # 切片保证不超过15个字符

    # 在文件名后加上"_copy"后缀
    new_pdf_name = new_pdf_base_name + "_copy" + extension

    # 构造新的PDF路径
    new_pdf_path = os.path.join(folder_path, new_pdf_name)

    # 复制文件到新路径
    shutil.copyfile(pdf_path, new_pdf_path)

    images_path = os.path.join(folder_path, results_folder, 'images')
    clip_path = os.path.join(folder_path, results_folder, 'clip')
    os.makedirs(f"{images_path}", exist_ok=True)
    os.makedirs(f"{clip_path}", exist_ok=True)
    images = convert_from_path(new_pdf_path)

    # 去掉 new_pdf_name 中的 '_copy.pdf' 部分
    image_copy_name = os.path.splitext(new_pdf_name)[0]

    # 然后，去掉结尾的 '_copy'
    if image_copy_name.endswith('_copy'):
        image_copy_name = image_copy_name[:-5]

    for i, image in enumerate(images):
        image.save(f"{images_path}\\{image_copy_name}_page{i+1}.jpg", "JPEG")
        print(f"Saved:{image_copy_name}_page{i+1}.jpg")
        image_cv = cv2.imread(f"{images_path}\\{image_copy_name}_page{i+1}.jpg")
        output = ocr.ocr(f"{images_path}\\{image_copy_name}_page{i+1}.jpg")[0]

        coords_up, coords_down = None, None

        for keyword in keywords:
            if keyword == keywords[0]:
                for item in output:
                    text_block = item[1]
                    text = text_block[0]
                    if keyword.lower() in text.lower():
                        index = text.lower().find(keyword.lower())
                        text_before_keyword = text[:index].strip()
                        break
                else:
                    text_before_keyword = str(i+1)
            else:
                coords = get_coords_of_keyword(output, keyword)
                if coords:
                    if keyword == keywords[1]:
                        coords_up = coords[2][1]
                    elif keyword == keywords[2]:
                        coords_down = coords[0][1]

        if coords_up is None or coords_down is None:
            continue
        else:
            y1, y2 = int(min(coords_up, coords_down)), int(max(coords_up, coords_down))
            cropped_image = image_cv[y1:y2, :]
            # 使用 text_before_keyword 作为文件名
            cv2.imwrite(f"{clip_path}\\{text_before_keyword}.jpg", cropped_image)

    # 删除复制的 PDF 文件
    os.remove(new_pdf_path)

def image_ocr(image_path, excel_folder):
    """
    对图像进行OCR识别，并将结果保存到Excel文件中。

    参数:
    image_path (str): 图像文件的路径。
    excel_folder (str): 保存Excel文件的文件夹路径。

    返回:
    无返回值，识别结果以Excel文件形式保存。
    """
    # 从image_path提取文件名（不含扩展名）
    number = os.path.splitext(os.path.basename(image_path))[0]
    image_cv = cv2.imread(image_path)
    image_height = image_cv.shape[0]
    image_width = image_cv.shape[1]
    output = ocr.ocr(image_path)[0]

    # 将output的内容分为boxes,texts,probabilities
    boxes = [line[0] for line in output]
    texts = [line[1][0] for line in output]
    probabilities = [line[1][1] for line in output]

    image_boxes = image_cv.copy()
    for box, text in zip(boxes, texts):
        cv2.rectangle(
            image_boxes,
            (int(box[0][0]), int(box[0][1])),
            (int(box[2][0]), int(box[2][1])),
            (0, 0, 255),
            1,
        )
        cv2.putText(
            image_boxes,
            text,
            (int(box[0][0]), int(box[0][1])),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (255, 0, 0),
            1,
        )

    # --------------------Reconstruction--------------------
    # Get Horizontal and Vertical Lines
    im = image_cv.copy()
    horiz_boxes = []
    vert_boxes = []

    for box in boxes:
        x_h, x_v = 0, int(box[0][0])
        y_h, y_v = int(box[0][1]), 0
        width_h, width_v = image_width, int(box[2][0] - box[0][0])
        height_h, height_v = int(box[2][1] - box[0][1]), image_height

        horiz_boxes.append([x_h, y_h, x_h + width_h, y_h + height_h])
        vert_boxes.append([x_v, y_v, x_v + width_v, y_v + height_v])

        cv2.rectangle(
            im, (x_h, y_h), (x_h + width_h, y_h + height_h), (0, 0, 255), 1
        )
        cv2.rectangle(
            im, (x_v, y_v), (x_v + width_v, y_v + height_v), (0, 255, 0), 1
        )

    # Non-Max Suppression
    horiz_out = tf.image.non_max_suppression(
        horiz_boxes,
        probabilities,
        max_output_size=1000,
        iou_threshold=0.1,
        score_threshold=float("-inf"),
        name=None,
    )
    horiz_lines = np.sort(np.array(horiz_out))

    im_nms = image_cv.copy()
    for val in horiz_lines:
        cv2.rectangle(
            im_nms,
            (int(horiz_boxes[val][0]), int(horiz_boxes[val][1])),
            (int(horiz_boxes[val][2]), int(horiz_boxes[val][3])),
            (0, 0, 255),
            1,
        )

    vert_out = tf.image.non_max_suppression(
        vert_boxes,
        probabilities,
        max_output_size=1000,
        iou_threshold=0.05,
        score_threshold=float("-inf"),
        name=None,
    )
    vert_lines = np.sort(np.array(vert_out))

    for val in vert_lines:
        cv2.rectangle(
            im_nms,
            (int(vert_boxes[val][0]), int(vert_boxes[val][1])),
            (int(vert_boxes[val][2]), int(vert_boxes[val][3])),
            (255, 0, 0),
            1,
        )

    # Convert to EXCEL
    out_array = [
        ["" for i in range(len(vert_lines))] for j in range(len(horiz_lines))
    ]

    unordered_boxes = []

    for i in vert_lines:
        unordered_boxes.append(vert_boxes[i][0])
    ordered_boxes = np.argsort(unordered_boxes)

    def intersection(box_1, box_2):
        return [box_2[0], box_1[1], box_2[2], box_1[3]]

    def iou(box_1, box_2):
        x_1 = max(box_1[0], box_2[0])
        y_1 = max(box_1[1], box_2[1])
        x_2 = min(box_1[2], box_2[2])
        y_2 = min(box_1[3], box_2[3])

        inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
        if inter == 0:
            return 0

        box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
        box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

        return inter / float(box_1_area + box_2_area - inter)

    for i in range(len(horiz_lines)):
        for j in range(len(vert_lines)):
            resultant = intersection(
                horiz_boxes[horiz_lines[i]],
                vert_boxes[vert_lines[ordered_boxes[j]]],
            )

            for b in range(len(boxes)):
                the_box = [
                    boxes[b][0][0],
                    boxes[b][0][1],
                    boxes[b][2][0],
                    boxes[b][2][1],
                ]
                if iou(resultant, the_box) > 0.1:
                    out_array[i][j] = texts[b]

    out_array = np.array(out_array)
    df = pd.DataFrame(out_array)

    excel_file_path = os.path.join(excel_folder, f"{number}.xlsx")
    df.to_excel(excel_file_path, index=False)

    # 加载 Excel 文件
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    # 删除第一列
    sheet.delete_rows(1)

    # 保存修改后的文件
    workbook.save(excel_file_path)

def process_images_in_folder(folder_path, excel_folder):
    """
    对文件夹中的所有图像文件进行OCR识别，并将结果保存到Excel文件中。

    参数:
    folder_path (str): 包含图像文件的文件夹路径。
    excel_folder (str): 保存Excel文件的文件夹路径。

    返回:
    无返回值，为每个图像生成对应的Excel文件。
    """
    # 获取folder_path下所有.jpg文件
    for image_path in glob.glob(os.path.join(folder_path, "*.jpg")):
        image_ocr(image_path, excel_folder)
        print(f"Processed and created Excel for {image_path}")

def merge_excels_to_sheets(source_folder, output_excel_path):
    """
    将一个文件夹中的所有Excel文件合并为一个Excel文件，每个原文件为一个工作表。

    参数:
    source_folder (str): 包含源Excel文件的文件夹路径。
    output_excel_path (str): 输出的Excel文件路径。

    返回:
    无返回值，合并后的Excel文件保存到output_excel_path。
    """
    # 创建一个Excel writer
    writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')

    # 遍历source_folder中的所有Excel文件
    for excel_file in glob.glob(os.path.join(source_folder, "*.xlsx")):
        # 使用原Excel文件名（不含扩展名）作为sheet名
        sheet_name = os.path.splitext(os.path.basename(excel_file))[0]

        # 读取Excel文件到DataFrame
        df = pd.read_excel(excel_file)

        # 将DataFrame写入新Excel文件的一个sheet中
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 保存新的Excel文件
    writer.close()

def setup_environment(base_path):
    """
    设置程序运行的环境，创建必要的文件夹。

    参数:
    base_path (str): 基础路径，所有输出文件夹将在该路径下创建。

    返回:
    无返回值，创建必要的文件夹结构。
    """
    # 防止系统崩溃
    # os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    os.makedirs(os.path.join(base_path, "Excel_union"), exist_ok=True)
    os.makedirs(os.path.join(base_path, "Data"), exist_ok=True)

# 重命名文件以避免报错
def rename_pdf_files(pdf_folder):
    """
    重命名PDF文件，保留字母和数字，避免文件名导致读取错误。

    参数:
    pdf_folder (str): PDF文件所在的文件夹路径。

    返回:
    无返回值，文件重命名操作直接在文件系统中完成。
    """
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    for pdf_path in pdf_files:
        pdf_name_with_extension = os.path.basename(pdf_path)
        # 过滤文件名，保留字母、数字（空格可能会导致读取时报错）
        filtered_name = re.sub(r'[^a-zA-Z0-9]', '', os.path.splitext(pdf_name_with_extension)[0])
        filtered_name = re.sub(r'\s+', ' ', filtered_name)
        # 删除文件名末尾的空格
        filtered_name = filtered_name.rstrip()

        # 如果过滤后的文件名超过20个字符，截断它以保持在20个字符以内
        if len(filtered_name) > 20:
            filtered_name = filtered_name[:20]

        # 构造新的PDF路径
        filtered_pdf_path = os.path.join(pdf_folder, f'{filtered_name}.pdf')

        # 重命名PDF文件
        if pdf_path != filtered_pdf_path:
            shutil.move(pdf_path, filtered_pdf_path)

def copy_pdf_folder(original_folder, backup_folder):
    """
    复制整个PDF文件夹到备份位置。

    参数:
    original_folder (str): 源PDF文件夹路径。
    backup_folder (str): 备份PDF文件夹路径。

    返回:
    无返回值，文件夹复制操作直接在文件系统中完成。
    """
    if not os.path.exists(backup_folder):
        shutil.copytree(original_folder, backup_folder)

def restore_original_pdf_folder(backup_folder, original_folder):
    """
    使用备份文件夹恢复原始的PDF文件夹。

    参数:
    backup_folder (str): 备份文件夹的路径。
    original_folder (str): 原始文件夹的路径，将被还原。

    返回:
    无返回值，将备份文件夹还原为原始文件夹。
    """
    if os.path.exists(original_folder):
        shutil.rmtree(original_folder)
    shutil.move(backup_folder, original_folder)

def intersection(box_1, box_2):
    """
    计算两个矩形框的交集。

    参数:
    box_1 (list of int): 第一个矩形框的坐标 [x1, y1, x2, y2]。
    box_2 (list of int): 第二个矩形框的坐标 [x1, y1, x2, y2]。

    返回:
    list: 表示交集区域的矩形框坐标 [x1, y1, x2, y2]。
    """
    return [box_2[0], box_1[1], box_2[2], box_1[3]]

# 计算iou(相交部分\\总的面积) 用于ocr
def iou(box_1, box_2):
    """
    计算两个矩形框的交并比（Intersection over Union, IoU）。

    参数:
    box_1 (list of int): 第一个矩形框的坐标 [x1, y1, x2, y2]。
    box_2 (list of int): 第二个矩形框的坐标 [x1, y1, x2, y2]。

    返回:
    float: 两个矩形框的交并比（IoU）值，范围为 0 到 1。
    """
    x_1 = max(box_1[0], box_2[0])
    y_1 = max(box_1[1], box_2[1])
    x_2 = min(box_1[2], box_2[2])
    y_2 = min(box_1[3], box_2[3])

    inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
    if inter == 0:
        return 0

    box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
    box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

    return inter / float(box_1_area + box_2_area - inter)

def copy_sheet(source_sheet, target_sheet):
    """
    将源工作表的内容复制到目标工作表。

    参数:
    source_sheet (openpyxl.worksheet.worksheet.Worksheet): 源工作表对象。
    target_sheet (openpyxl.worksheet.worksheet.Worksheet): 目标工作表对象。

    返回:
    无返回值，内容被复制到目标工作表。
    """
    for row in source_sheet:
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

# 为每个PDF创建相应文件夹
def img_initialization(pages_dir, results_dir, images, pdf_name_without_extension, table_ok, figure_ok):
    """
    初始化图像处理的目录结构，并保存PDF每页的图像。

    参数:
    pages_dir (str): 保存处理后图像的文件夹路径。
    results_dir (str): 保存结果文件的文件夹路径。
    images (list of PIL.Image): PDF转换后的图像列表。
    pdf_name_without_extension (str): PDF文件名（不含扩展名）。
    table_ok (bool): 是否提取表格信息。
    figure_ok (bool): 是否提取图像信息。

    返回:
    无返回值，初始化的文件夹结构和图像被保存。
    """
    os.makedirs(pages_dir, exist_ok=True)
    os.makedirs(results_dir, exist_ok=True)
    if figure_ok:
        os.makedirs(f"{results_dir}\\figure", exist_ok=True)
    if table_ok:
        os.makedirs(f"{results_dir}\\table", exist_ok=True)
        os.makedirs(f"{results_dir}\\detections", exist_ok=True)
        os.makedirs(f"{results_dir}\\horiz_vert", exist_ok=True)
        os.makedirs(f"{results_dir}\\im_nms", exist_ok=True)
        os.makedirs(f"{results_dir}\\excel", exist_ok=True)

    for i in range(len(images)):
        images[i].save(
            f"{pages_dir}\\{pdf_name_without_extension}_page" + str(i) + ".jpg", "JPEG"
        )
        print(f"{pages_dir}\\{pdf_name_without_extension}_page" + str(i) + ".jpg")
# Paddle模型解析PDF图片,分为0: "Text", 1: "Title", 2: "List", 3: "Table", 4: "Figure"
def pdf_detection(pages_dir, pdf_name_without_extension, page_num, model_threshold):
    """
    对PDF页面进行对象检测，提取文本、表格和图像区域。

    参数:
    pages_dir (str): 保存页面图像的文件夹路径。
    pdf_name_without_extension (str): PDF文件名（不含扩展名）。
    page_num (int): 当前处理的页面编号。
    model_threshold (float): 模型检测的可信度阈值。

    返回:
    layout (list): 检测到的对象布局信息，包括文本、表格和图像。
    """
    image_path = f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg"
    print(image_path)

    image = cv2.imread(image_path)
    image = image[..., ::-1]
    print(f"正在处理文件：{pdf_name_without_extension}, 页码：{page_num}")

    # load model
    model = lp.PaddleDetectionLayoutModel(
        config_path="lp://PubLayNet/ppyolov2_r50vd_dcn_365e_publaynet/config",
        # 可信度达到0.5认为是某一类
        threshold = model_threshold,
        label_map={0: "Text", 1: "Title", 2: "List", 3: "Table", 4: "Figure"},
        enforce_cpu=False,
        enable_mkldnn=True,
    )  # math kernel library
    # detect
    layout = model.detect(image)
    return layout

# 合并一个PDF对应的所有生成的xlsx文件
def create_excel(results_dir, pdf_name_without_extension, base_path):
    """
    将一个PDF对应的所有生成的Excel文件合并为一个。

    参数:
    results_dir (str): 存储中间Excel文件的目录。
    pdf_name_without_extension (str): PDF文件名（不含扩展名）。
    base_path (str): 基础路径，保存最终的合并Excel文件。

    返回:
    无返回值，合并后的Excel文件保存到base_path路径下。
    """
    wb = Workbook()
    excel_dir = f"{results_dir}\\excel"
    final_excel_path = os.path.join(base_path, f"Excel_union\\{pdf_name_without_extension}.xlsx")

    excel_files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx")]

    # 如果目录中没有Excel文件，添加一个默认的工作表
    if not excel_files:
        wb.create_sheet(title="Sheet1")
    else:
        # 如果有Excel文件，则移除默认添加的工作表
        wb.remove(wb.active)

    for excel_file in excel_files:
        file_path = os.path.join(excel_dir, excel_file)
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active

        # 使用文件名（不包括扩展名）作为工作表名称
        sheet_name = "_".join(excel_file.split("_")[1:])
        sheet_name = os.path.splitext(sheet_name)[0]

        # 限制工作表名称为前25个字符
        actual_sheet_name = sheet_name[:20]

        # 创建新的工作表，名称为实际的工作表名称
        target_sheet = wb.create_sheet(title=actual_sheet_name)

        # 复制内容到新工作表
        copy_sheet(source_sheet, target_sheet)

        # 找到目标工作表中最后一行的行号
        last_row = target_sheet.max_row

        # 在所有复制的内容之后的第二行插入文件名（不包括扩展名）
        target_sheet.cell(row=last_row + 2, column=1, value=sheet_name)

    # 保存最终的工作簿
    wb.save(final_excel_path)


def process_pdf_file(pdf_path, base_path, table_ok, figure_ok, model_threshold):
    """
    处理一个PDF文件，提取表格和图像，并生成对应的Excel文件。

    参数:
    pdf_path (str): PDF文件的路径。
    base_path (str): 程序的基础路径，用于保存处理结果。
    table_ok (bool): 是否提取表格信息。
    figure_ok (bool): 是否提取图像信息。
    model_threshold (float): 模型检测的可信度阈值。

    返回:
    无返回值，处理结果保存在指定路径下。
    """
    pdf_name_with_extension = os.path.basename(pdf_path)
    pdf_name_without_extension = os.path.splitext(pdf_name_with_extension)[0]

    images = convert_from_path(pdf_path)

    print("image finish")

    pages_dir = os.path.join(base_path, f"Data\\{pdf_name_without_extension}\\{pdf_name_without_extension}_pages")
    results_dir = (
        os.path.join(base_path, f"Data\\{pdf_name_without_extension}\\{pdf_name_without_extension}_results")
    )

    img_initialization(pages_dir, results_dir, images, pdf_name_without_extension, table_ok, figure_ok)

    # --------------------Table Extraction--------------------
    # 遍历每一页，定位表格与图片
    for page_idx in range(len(images)):
        page_num = f"page{page_idx}"

        layout = pdf_detection(pages_dir, pdf_name_without_extension, page_num, model_threshold)

        table_loc = []
        figure_loc = []

        # 获得表格框坐标（适当扩展）
        for l in layout:
            # 表格定位
            if table_ok:
                if l.type == "Table":
                    x_1 = int(l.block.x_1) - 4
                    y_1 = math.ceil(l.block.y_1) + 4
                    x_2 = math.ceil(l.block.x_2) + 4
                    y_2 = int(l.block.y_2) - 4
                    table_loc.append((x_1, y_1, x_2, y_2))
            # 图片定位
            if figure_ok:
                if l.type == 'Figure':
                    x_3 = int(l.block.x_1) - 4
                    y_3 = math.ceil(l.block.y_1) + 4
                    x_4 = math.ceil(l.block.x_2) + 4
                    y_4 = int(l.block.y_2) - 4
                    figure_loc.append((x_3, y_3, x_4, y_4))

        if figure_ok:
            for idx, cor in enumerate(figure_loc):
                x_3, y_3, x_4, y_4 = cor

                im = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\figure\\{pdf_name_without_extension}_{page_num}_figure_{idx}.jpg",
                    im[y_3:y_4, x_3:x_4],
                )
        if table_ok:
            for idx, cor in enumerate(table_loc):
                x_1, y_1, x_2, y_2 = cor
                combined_text = None
                im_text = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_text_{idx}.jpg",
                    im_text[0:int(y_1), int(x_1):int(x_2)],
                )
                ocr = PaddleOCR(lang="en")
                image_path = f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_text_{idx}.jpg"
                image_cv = cv2.imread(image_path)
                image_height = image_cv.shape[0]
                image_width = image_cv.shape[1]
                output_table = ocr.ocr(image_path)[0]
                if output_table and output_table[0]:  # 确保output_table不是None且包含识别结果
                    texts_table = [line_table[1][0] for line_table in output_table]
                    texts_table_lower = [text.lower() for text in texts_table]

                    # Initialize a variable to hold the index of 'table' if found
                    table_index = None

                    # Iterate over texts_table_lower in reverse to find 'table'
                    for i, text in enumerate(reversed(texts_table_lower)):
                        if 'table' in text:
                            # Calculate the correct index in the original list
                            table_index = len(texts_table_lower) - 1 - i
                            break

                    # Check if 'table' was found
                    if table_index is not None:
                        # Extract and print the texts from 'table' onwards
                        result_texts = texts_table[table_index:]
                        combined_text = ' '.join(result_texts)
                        combined_text = combined_text[:140]
                    else:
                        combined_text = None

                    if combined_text is not None:
                        # Replace all characters not a letter, a number, or a space with '_'
                        combined_text = re.sub(r'[^a-zA-Z0-9\s]', '_', combined_text)

                im = cv2.imread(f"{pages_dir}\\{pdf_name_without_extension}_{page_num}.jpg")
                cv2.imwrite(
                    f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_table_{idx}.jpg",
                    im[y_1:y_2, x_1:x_2],
                )

                # --------------------Text Detection and Recognition--------------------
                # OCR部分，可独立
                ocr = PaddleOCR(lang="en")
                image_path = f"{results_dir}\\table\\{pdf_name_without_extension}_{page_num}_table_{idx}.jpg"
                image_cv = cv2.imread(image_path)
                image_height = image_cv.shape[0]
                image_width = image_cv.shape[1]
                output = ocr.ocr(image_path)[0]

                # 将output的内容分为boxes,texts,probabilities
                boxes = [line[0] for line in output]
                texts = [line[1][0] for line in output]
                probabilities = [line[1][1] for line in output]

                image_boxes = image_cv.copy()
                for box, text in zip(boxes, texts):
                    cv2.rectangle(
                        image_boxes,
                        (int(box[0][0]), int(box[0][1])),
                        (int(box[2][0]), int(box[2][1])),
                        (0, 0, 255),
                        1,
                    )
                    cv2.putText(
                        image_boxes,
                        text,
                        (int(box[0][0]), int(box[0][1])),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.6,
                        (255, 0, 0),
                        1,
                    )
                cv2.imwrite(
                    f"{results_dir}\\detections\\{pdf_name_without_extension}_{page_num}_detections_{idx}.jpg",
                    image_boxes,
                )

                # --------------------Reconstruction--------------------
                # Get Horizontal and Vertical Lines
                im = image_cv.copy()
                horiz_boxes = []
                vert_boxes = []

                for box in boxes:
                    x_h, x_v = 0, int(box[0][0])
                    y_h, y_v = int(box[0][1]), 0
                    width_h, width_v = image_width, int(box[2][0] - box[0][0])
                    height_h, height_v = int(box[2][1] - box[0][1]), image_height

                    horiz_boxes.append([x_h, y_h, x_h + width_h, y_h + height_h])
                    vert_boxes.append([x_v, y_v, x_v + width_v, y_v + height_v])

                    cv2.rectangle(
                        im, (x_h, y_h), (x_h + width_h, y_h + height_h), (0, 0, 255), 1
                    )
                    cv2.rectangle(
                        im, (x_v, y_v), (x_v + width_v, y_v + height_v), (0, 255, 0), 1
                    )
                cv2.imwrite(
                    f"{results_dir}\\horiz_vert\\{pdf_name_without_extension}_{page_num}_horiz_vert_{idx}.jpg",
                    im,
                )

                # Non-Max Suppression
                horiz_out = tf.image.non_max_suppression(
                    horiz_boxes,
                    probabilities,
                    max_output_size=1000,
                    iou_threshold=0.1,
                    score_threshold=float("-inf"),
                    name=None,
                )
                horiz_lines = np.sort(np.array(horiz_out))

                im_nms = image_cv.copy()
                for val in horiz_lines:
                    cv2.rectangle(
                        im_nms,
                        (int(horiz_boxes[val][0]), int(horiz_boxes[val][1])),
                        (int(horiz_boxes[val][2]), int(horiz_boxes[val][3])),
                        (0, 0, 255),
                        1,
                    )

                vert_out = tf.image.non_max_suppression(
                    vert_boxes,
                    probabilities,
                    max_output_size=1000,
                    iou_threshold=0.05,
                    score_threshold=float("-inf"),
                    name=None,
                )
                vert_lines = np.sort(np.array(vert_out))

                for val in vert_lines:
                    cv2.rectangle(
                        im_nms,
                        (int(vert_boxes[val][0]), int(vert_boxes[val][1])),
                        (int(vert_boxes[val][2]), int(vert_boxes[val][3])),
                        (255, 0, 0),
                        1,
                    )
                cv2.imwrite(
                    f"{results_dir}\\im_nms\\{pdf_name_without_extension}_{page_num}_im_nms_{idx}.jpg",
                    im_nms,
                )

                # Convert to EXCEL
                out_array = [
                    ["" for i in range(len(vert_lines))] for j in range(len(horiz_lines))
                ]

                unordered_boxes = []

                for i in vert_lines:
                    unordered_boxes.append(vert_boxes[i][0])
                ordered_boxes = np.argsort(unordered_boxes)

                def intersection(box_1, box_2):
                    return [box_2[0], box_1[1], box_2[2], box_1[3]]

                def iou(box_1, box_2):
                    x_1 = max(box_1[0], box_2[0])
                    y_1 = max(box_1[1], box_2[1])
                    x_2 = min(box_1[2], box_2[2])
                    y_2 = min(box_1[3], box_2[3])

                    inter = abs(max((x_2 - x_1, 0)) * max((y_2 - y_1), 0))
                    if inter == 0:
                        return 0

                    box_1_area = abs((box_1[2] - box_1[0]) * (box_1[3] - box_1[1]))
                    box_2_area = abs((box_2[2] - box_2[0]) * (box_2[3] - box_2[1]))

                    return inter / float(box_1_area + box_2_area - inter)

                for i in range(len(horiz_lines)):
                    for j in range(len(vert_lines)):
                        resultant = intersection(
                            horiz_boxes[horiz_lines[i]],
                            vert_boxes[vert_lines[ordered_boxes[j]]],
                        )

                        for b in range(len(boxes)):
                            the_box = [
                                boxes[b][0][0],
                                boxes[b][0][1],
                                boxes[b][2][0],
                                boxes[b][2][1],
                            ]
                            if iou(resultant, the_box) > 0.1:
                                out_array[i][j] = texts[b]

                out_array = np.array(out_array)
                df = pd.DataFrame(out_array)
                df.to_excel(
                    f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx",
                    index=False,
                )

                # 加载 Excel 文件
                workbook = load_workbook(
                    filename=f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx"
                )
                sheet = workbook.active

                # 删除第一列
                sheet.delete_rows(1)

                # 保存修改后的文件
                workbook.save(
                    f"{results_dir}\\excel\\{pdf_name_without_extension}_{page_num}_{idx}_{combined_text}.xlsx"
                )
    if table_ok:
        create_excel(results_dir, pdf_name_without_extension, base_path)


def main(base_path, table_ok, figure_ok, model_threshold, mode, base_folder, keywords):
    """
    主函数，控制程序的流程，处理PDF文件，提取表格或关键词对应的图像区域。

    参数:
    base_path (str): 基础路径，存储处理结果。
    table_ok (bool): 是否提取表格信息。
    figure_ok (bool): 是否提取图像信息。
    model_threshold (float): 模型检测的可信度阈值。
    mode (str): 处理模式，'normal' 或 'keyword'。
    base_folder (str): PDF文件所在的基础文件夹路径。
    keywords (list of str): 需要查找的关键词列表（仅在 'keyword' 模式下有效）。

    返回:
    无返回值，完成指定模式下的PDF文件处理。
    """
    print(mode)
    if mode == 'normal':
        backup_folder = os.path.join(base_path, "Document_Copy")

        setup_environment(base_path)

        # 创建Document文件夹（如果已有则清空）
        pdf_folder = os.path.join(base_path, "Document")
        if os.path.exists(pdf_folder):
            shutil.rmtree(pdf_folder)
        os.makedirs(pdf_folder)

        # 将base_path中的PDF文件复制到Document文件夹中
        for pdf_file in glob.glob(os.path.join(base_path, "*.pdf")):
            shutil.copy(pdf_file, pdf_folder)

        # 复制PDF文件夹
        copy_pdf_folder(pdf_folder, backup_folder)

        rename_pdf_files(pdf_folder)

        # 获取文件夹内所有的PDF文件
        pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))

        # 遍历裁剪&OCR
        for pdf_path in pdf_files:
            process_pdf_file(pdf_path, base_path, table_ok, figure_ok, model_threshold)

        restore_original_pdf_folder(backup_folder, pdf_folder)

        # 删除Document文件夹
        shutil.rmtree(pdf_folder)

    if mode == 'keyword':
        for pdf_file in os.listdir(base_folder):
            if pdf_file.endswith('.pdf'):
                pdf_path = os.path.join(base_folder, pdf_file)

                pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
                results_folder = f'{pdf_name}_results'
                ocr_and_crop(pdf_path, keywords, results_folder)

                base_path = os.path.join(base_folder, results_folder)
                clip_folder_path = os.path.join(base_path, 'clip')
                excel_path = os.path.join(base_path, "union.xlsx")
                excel_folder = os.path.join(base_path, "excel")
                if not os.path.exists(excel_folder):
                    os.makedirs(excel_folder)

                if os.listdir(clip_folder_path):
                    # 如果clip文件夹不为空，则执行下列函数
                    process_images_in_folder(clip_folder_path, excel_folder)
                    merge_excels_to_sheets(excel_folder, excel_path)

if __name__ == "__main__":
    # 可使用脚本传参，或调整默认参数运行
    parser = argparse.ArgumentParser()
    # pdf存放路径 推荐使用\\
    parser.add_argument("--base_path", type=str, default=r"F:\\OCR\\TEST")
    # 是否提取表格
    parser.add_argument("--table_ok", type=bool, default=True)
    # 是否提取图像
    parser.add_argument("--figure_ok", type=bool, default=True)
    # 模型检测的可信度阈值
    parser.add_argument("--model_threshold", type=float, default=0.5)
    # 处理模式 （一般论文使用normal，关键词功能并不完善）
    parser.add_argument("--mode", type=str, default='normal')
    # 关键词模式下存放路径
    parser.add_argument('--base_folder', type=str, default=r"base_folder")
    # 关键词列表
    parser.add_argument('--keywords', type=str, nargs='+', default=["Keywords"])

    args = parser.parse_args()

    main(args.base_path, args.table_ok, args.figure_ok, args.model_threshold, args.mode, args.base_folder, args.keywords)

